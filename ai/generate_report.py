"""
Generate Excel Report dari hasil training YOLOv8
Isi report:
  Sheet 1 - Ringkasan Model       : info model, best metrics, inference time
  Sheet 2 - Metrik Per Epoch      : tabel lengkap semua epoch
  Sheet 3 - Metrik Per Kelas      : precision/recall/mAP per golongan kendaraan
  Sheet 4 - Grafik Training       : embed gambar results.png, confusion matrix, dll
"""

import csv
import os
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from ultralytics import YOLO

# ── CONFIG ────────────────────────────────────────────────────────────────────
RESULTS_DIR = Path(r"C:\Users\aryas\OneDrive\Documents\Binus Univ\Semester 2\AI\Vehicle_Classification\vehicle_cls_v1_results")
DATA_YAML   = Path(r"C:\Users\aryas\OneDrive\Documents\Binus Univ\Semester 2\AI\Vehicle_Classification\ai\dataset\data.yaml")
OUTPUT_FILE = RESULTS_DIR / "Training_Report_YOLOv8.xlsx"

CLASS_NAMES = ["GOL I", "GOL II", "GOL III", "GOL IV", "GOL V"]

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
BLUE_DARK   = "1E3A6E"
BLUE_MID    = "2D5AA0"
BLUE_LIGHT  = "DBEAFE"
YELLOW      = "FCD34D"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F8FAFC"
GRAY_BORDER = "CBD5E1"
GREEN       = "16A34A"
RED         = "DC2626"

def hdr_font(bold=True, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def body_font(bold=False, color="1E293B", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color=GRAY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def pct(val):
    return f"{float(val)*100:.2f}%"

def fmt(val, decimals=4):
    return f"{float(val):.{decimals}f}"

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header_row(ws, row, cols, bg=BLUE_DARK):
    for c, (col, text, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font      = hdr_font()
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
        set_col_width(ws, c, width)

def write_data_row(ws, row, values, alt=False):
    bg = GRAY_LIGHT if alt else WHITE
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = body_font()
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = thin_border()

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
print("📂 Membaca results.csv ...")
csv_path = RESULTS_DIR / "results.csv"
epochs_data = []
with open(csv_path, newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        epochs_data.append({k.strip(): v.strip() for k, v in row.items()})

best_map50    = max(float(r["metrics/mAP50(B)"])    for r in epochs_data)
best_map5095  = max(float(r["metrics/mAP50-95(B)"]) for r in epochs_data)
best_prec     = max(float(r["metrics/precision(B)"]) for r in epochs_data)
best_recall   = max(float(r["metrics/recall(B)"])   for r in epochs_data)
best_epoch_row = max(epochs_data, key=lambda r: float(r["metrics/mAP50(B)"]))
best_epoch    = int(best_epoch_row["epoch"])
total_epochs  = len(epochs_data)

# ── INFERENCE TIME via YOLO ───────────────────────────────────────────────────
print("⏱️  Mengukur inference time ...")
model_path = RESULTS_DIR / "weights" / "best.pt"
model = YOLO(str(model_path))

# Dummy inference untuk warmup
import numpy as np
dummy = np.zeros((640, 640, 3), dtype=np.uint8)
for _ in range(3):
    model.predict(dummy, verbose=False)

# Ukur 10x
times = []
for _ in range(10):
    t0 = time.perf_counter()
    model.predict(dummy, verbose=False)
    times.append((time.perf_counter() - t0) * 1000)

inf_mean = sum(times) / len(times)
inf_min  = min(times)
inf_max  = max(times)

# ── EVALUASI PER KELAS ────────────────────────────────────────────────────────
print("📊 Evaluasi per kelas ...")
metrics = model.val(data=str(DATA_YAML), verbose=False)

per_class = []
for i, name in enumerate(CLASS_NAMES):
    ap50 = float(metrics.box.ap50[i]) if i < len(metrics.box.ap50) else 0
    ap   = float(metrics.box.ap[i])   if i < len(metrics.box.ap)   else 0
    per_class.append({
        "name": name,
        "ap50": ap50,
        "ap":   ap,
    })

# ── CREATE WORKBOOK ───────────────────────────────────────────────────────────
print("📝 Membuat Excel ...")
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — RINGKASAN MODEL
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Ringkasan Model"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 32
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 22
ws1.row_dimensions[1].height = 10

# Title
ws1.merge_cells("A2:D2")
t = ws1["A2"]
t.value     = "LAPORAN HASIL TRAINING — YOLOv8n"
t.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
t.fill      = fill(BLUE_DARK)
t.alignment = center()
ws1.row_dimensions[2].height = 36

ws1.merge_cells("A3:D3")
s = ws1["A3"]
s.value     = "Klasifikasi Golongan Kendaraan Jalan Tol"
s.font      = Font(bold=False, size=11, color=WHITE, name="Calibri")
s.fill      = fill(BLUE_MID)
s.alignment = center()
ws1.row_dimensions[3].height = 22

# ── Info Model ──
def section_title(ws, row, text, span="A:D"):
    ws.merge_cells(f"{span[0]}{row}:{span[2]}{row}")
    c = ws[f"{span[0]}{row}"]
    c.value     = text
    c.font      = Font(bold=True, size=11, color=BLUE_DARK, name="Calibri")
    c.fill      = fill(BLUE_LIGHT)
    c.alignment = left()
    c.border    = thin_border()
    ws.row_dimensions[row].height = 20

def kv_row(ws, row, label, value, value2="", value3=""):
    cells = [label, value, value2, value3]
    for c, val in enumerate(cells, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.border    = thin_border()
        cell.alignment = center() if c > 1 else left()
        cell.fill      = fill(GRAY_LIGHT if row % 2 == 0 else WHITE)
        if c == 1:
            cell.font = body_font(bold=True)
        else:
            cell.font = body_font()
    ws.row_dimensions[row].height = 18

section_title(ws1, 5, "  ℹ️  Informasi Model")
ws1.cell(row=6, column=1, value="Parameter").font  = hdr_font(color=BLUE_DARK)
ws1.cell(row=6, column=2, value="Nilai").font       = hdr_font(color=BLUE_DARK)
for c in range(1, 5):
    ws1.cell(row=6, column=c).fill      = fill(BLUE_LIGHT)
    ws1.cell(row=6, column=c).border    = thin_border()
    ws1.cell(row=6, column=c).alignment = center()

info_rows = [
    ("Nama Model",          "YOLOv8n (Nano)"),
    ("Task",                "Object Detection"),
    ("Dataset",             "Golongan Kendaraan Jalan Tol"),
    ("Jumlah Kelas",        "5 (GOL I – GOL V)"),
    ("Total Epoch",         str(total_epochs)),
    ("Best Epoch",          str(best_epoch)),
    ("Ukuran Input",        "640 × 640 px"),
    ("Batch Size",          "16"),
    ("Optimizer",           "SGD (default YOLOv8)"),
    ("Pretrained Weights",  "yolov8n.pt (COCO)"),
]
for i, (k, v) in enumerate(info_rows):
    kv_row(ws1, 7 + i, k, v)

# ── Metrik Terbaik ──
r = 18
section_title(ws1, r, "  📈  Metrik Terbaik (Best Epoch)")
r += 1
headers = [("Metrik", 32), ("Nilai", 22), ("Persentase", 22), ("Keterangan", 22)]
for c, (h, _) in enumerate(headers, 1):
    cell = ws1.cell(row=r, column=c, value=h)
    cell.font = hdr_font(color=BLUE_DARK); cell.fill = fill(BLUE_LIGHT)
    cell.alignment = center(); cell.border = thin_border()
r += 1

metric_rows = [
    ("Precision (Presisi)",          best_prec,   pct(best_prec),   "Akurasi prediksi positif"),
    ("Recall (Daya Ingat)",          best_recall, pct(best_recall), "Kemampuan menemukan semua objek"),
    ("mAP@0.50",                     best_map50,  pct(best_map50),  "Standar utama deteksi objek"),
    ("mAP@0.50:0.95",                best_map5095,pct(best_map5095),"Metrik ketat multi-threshold"),
    ("Inference Time (rata-rata)",   inf_mean,    f"{inf_mean:.1f} ms", "Waktu deteksi per gambar"),
    ("Inference Time (minimum)",     inf_min,     f"{inf_min:.1f} ms",  "Waktu tercepat"),
    ("Inference Time (maksimum)",    inf_max,     f"{inf_max:.1f} ms",  "Waktu terlambat"),
]
for i, (label, raw, display, note) in enumerate(metric_rows):
    alt = i % 2 == 0
    bg  = GRAY_LIGHT if alt else WHITE
    row_vals = [label, fmt(raw), display, note]
    for c, val in enumerate(row_vals, 1):
        cell = ws1.cell(row=r + i, column=c, value=val)
        cell.fill = fill(bg); cell.border = thin_border()
        cell.alignment = center() if c > 1 else left()
        cell.font = body_font(bold=(c == 1))
        # Warnai nilai mAP hijau jika > 90%
        if c == 3 and "mAP" in label and raw >= 0.9:
            cell.font = Font(bold=True, color=GREEN, size=10, name="Calibri")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — METRIK PER EPOCH
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Metrik Per Epoch")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:N1")
t2 = ws2["A1"]
t2.value     = "Metrik Training Per Epoch — YOLOv8n"
t2.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
t2.fill      = fill(BLUE_DARK)
t2.alignment = center()
ws2.row_dimensions[1].height = 30

cols2 = [
    ("A", "Epoch",          7),
    ("B", "Waktu (s)",      10),
    ("C", "Loss Box",       11),
    ("D", "Loss Cls",       11),
    ("E", "Loss DFL",       11),
    ("F", "Precision",      12),
    ("G", "Recall",         12),
    ("H", "mAP@0.50",       12),
    ("I", "mAP@0.50:0.95",  14),
    ("J", "Val Box Loss",   12),
    ("K", "Val Cls Loss",   12),
    ("L", "Val DFL Loss",   12),
    ("M", "LR pg0",         10),
    ("N", "LR pg1",         10),
]
for c, (col, text, width) in enumerate(cols2, 1):
    cell = ws2.cell(row=2, column=c, value=text)
    cell.font = hdr_font(); cell.fill = fill(BLUE_DARK)
    cell.alignment = center(); cell.border = thin_border()
    ws2.column_dimensions[col].width = width
ws2.row_dimensions[2].height = 22

for i, row in enumerate(epochs_data):
    r = i + 3
    alt = i % 2 == 0
    vals = [
        int(row["epoch"]),
        float(row["time"]),
        float(row["train/box_loss"]),
        float(row["train/cls_loss"]),
        float(row["train/dfl_loss"]),
        float(row["metrics/precision(B)"]),
        float(row["metrics/recall(B)"]),
        float(row["metrics/mAP50(B)"]),
        float(row["metrics/mAP50-95(B)"]),
        float(row["val/box_loss"]),
        float(row["val/cls_loss"]),
        float(row["val/dfl_loss"]),
        float(row["lr/pg0"]),
        float(row["lr/pg1"]),
    ]
    bg = GRAY_LIGHT if alt else WHITE
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=round(val, 5) if isinstance(val, float) else val)
        cell.fill = fill(bg); cell.border = thin_border()
        cell.alignment = center(); cell.font = body_font()
        # Highlight best epoch
        if int(row["epoch"]) == best_epoch:
            cell.fill = fill("FEF9C3")
            cell.font = Font(bold=True, size=10, color="92400E", name="Calibri")

# Freeze header
ws2.freeze_panes = "A3"

# ── Line chart: mAP50 per epoch ──
chart = LineChart()
chart.title    = "mAP@0.50 per Epoch"
chart.style    = 10
chart.y_axis.title = "mAP@0.50"
chart.x_axis.title = "Epoch"
chart.width    = 18
chart.height   = 12

data_ref = Reference(ws2, min_col=8, min_row=2, max_row=2 + total_epochs)
chart.add_data(data_ref, titles_from_data=True)
chart.series[0].graphicalProperties.line.solidFill = BLUE_MID
chart.series[0].graphicalProperties.line.width     = 20000

cats = Reference(ws2, min_col=1, min_row=3, max_row=2 + total_epochs)
chart.set_categories(cats)
ws2.add_chart(chart, f"P3")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — METRIK PER KELAS
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Metrik Per Kelas")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value     = "Metrik Per Kelas Kendaraan — Best Model"
t3.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
t3.fill      = fill(BLUE_DARK)
t3.alignment = center()
ws3.row_dimensions[1].height = 30

# Overall row
overall_headers = [
    ("A", "Kelas",          20),
    ("B", "Deskripsi",      32),
    ("C", "AP@0.50",        14),
    ("D", "AP@0.50:0.95",   16),
    ("E", "AP@0.50 (%)",    14),
    ("F", "AP@0.50:0.95 (%)",16),
]
for c, (col, text, width) in enumerate(overall_headers, 1):
    cell = ws3.cell(row=2, column=c, value=text)
    cell.font = hdr_font(); cell.fill = fill(BLUE_DARK)
    cell.alignment = center(); cell.border = thin_border()
    ws3.column_dimensions[col].width = width
ws3.row_dimensions[2].height = 22

CLASS_DESC = [
    "Sedan / Jip / Pick-up / Bus",
    "Truk 2 Gandar",
    "Truk 3 Gandar",
    "Truk 4 Gandar",
    "Truk 5 Gandar atau lebih",
]
CLASS_COLORS = ["22C55E", "3B82F6", "F59E0B", "EF4444", "8B5CF6"]

for i, cls in enumerate(per_class):
    r = i + 3
    vals = [
        cls["name"],
        CLASS_DESC[i],
        round(cls["ap50"], 4),
        round(cls["ap"],   4),
        pct(cls["ap50"]),
        pct(cls["ap"]),
    ]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border    = thin_border()
        cell.alignment = center() if c > 2 else left()
        cell.font      = body_font(bold=(c == 1), color=CLASS_COLORS[i] if c == 1 else "1E293B")
        cell.fill      = fill(GRAY_LIGHT if i % 2 == 0 else WHITE)
    ws3.row_dimensions[r].height = 20

# Overall average row
r_avg = len(per_class) + 3
avg50   = sum(c["ap50"] for c in per_class) / len(per_class)
avg5095 = sum(c["ap"]   for c in per_class) / len(per_class)
avg_vals = ["RATA-RATA", "Semua Kelas", round(avg50,4), round(avg5095,4), pct(avg50), pct(avg5095)]
for c, val in enumerate(avg_vals, 1):
    cell = ws3.cell(row=r_avg, column=c, value=val)
    cell.font      = Font(bold=True, size=10, color=WHITE, name="Calibri")
    cell.fill      = fill(BLUE_MID)
    cell.alignment = center() if c > 2 else left()
    cell.border    = thin_border()
ws3.row_dimensions[r_avg].height = 22

# Overall mAP from model.val
r_map = r_avg + 2
section_title(ws3, r_map, "  📊  Metrik Global (model.val)", span="A:F")
r_map += 1
global_metrics = [
    ("mAP@0.50",      metrics.box.map50, pct(metrics.box.map50)),
    ("mAP@0.50:0.95", metrics.box.map,   pct(metrics.box.map)),
    ("Precision",     metrics.box.mp,    pct(metrics.box.mp)),
    ("Recall",        metrics.box.mr,    pct(metrics.box.mr)),
]
for c, h in enumerate(["Metrik", "Nilai", "Persentase", "", "", ""], 1):
    cell = ws3.cell(row=r_map, column=c, value=h)
    cell.font = hdr_font(color=BLUE_DARK); cell.fill = fill(BLUE_LIGHT)
    cell.alignment = center(); cell.border = thin_border()
r_map += 1
for i, (label, val, pct_str) in enumerate(global_metrics):
    for c, v in enumerate([label, round(val,4), pct_str, "", "", ""], 1):
        cell = ws3.cell(row=r_map + i, column=c, value=v)
        cell.fill = fill(GRAY_LIGHT if i%2==0 else WHITE)
        cell.border = thin_border(); cell.alignment = center() if c>1 else left()
        cell.font = body_font(bold=(c==1))
        if c == 3 and val >= 0.9:
            cell.font = Font(bold=True, color=GREEN, size=10, name="Calibri")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — GRAFIK & VISUALISASI
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Grafik & Visualisasi")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:J1")
t4 = ws4["A1"]
t4.value     = "Grafik Hasil Training — YOLOv8n"
t4.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
t4.fill      = fill(BLUE_DARK)
t4.alignment = center()
ws4.row_dimensions[1].height = 30

image_files = [
    ("results.png",                   "A3",  "Grafik Loss & Metrik per Epoch"),
    ("confusion_matrix_normalized.png","A32", "Confusion Matrix (Normalized)"),
    ("BoxPR_curve.png",               "L3",  "Precision-Recall Curve"),
    ("BoxF1_curve.png",               "L32", "F1-Confidence Curve"),
]

for fname, anchor, caption in image_files:
    fpath = RESULTS_DIR / fname
    if fpath.exists():
        # Caption
        col_letter = anchor[0]
        row_num    = int(anchor[1:]) - 1
        if row_num >= 1:
            cap_cell = ws4.cell(row=row_num, column=ord(col_letter)-64, value=caption)
            cap_cell.font      = Font(bold=True, size=10, color=BLUE_DARK, name="Calibri")
            cap_cell.alignment = left()
        img = XLImage(str(fpath))
        img.width  = 480
        img.height = 320
        ws4.add_image(img, anchor)
        print(f"  ✅ Embed: {fname}")
    else:
        print(f"  ⚠️  Tidak ditemukan: {fname}")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n✅ Report berhasil dibuat!")
print(f"📁 Lokasi: {OUTPUT_FILE}")
