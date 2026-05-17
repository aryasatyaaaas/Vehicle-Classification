import sys, openpyxl, copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

XLSX = r'vehicle_cls_v1_results\Training_Report_YOLOv8.xlsx'
wb = openpyxl.load_workbook(XLSX)
ws = wb['Metrik Per Kelas']

# Data per kelas dari hasil YOLO val
# Urutan: GOL I, GOL II, GOL III, GOL IV, GOL V, RATA-RATA
precision_per_class = [0.9152, 0.8962, 0.7185, 1.0000, 0.9230, 0.8906]
recall_per_class    = [0.9211, 0.8718, 1.0000, 0.8619, 0.8560, 0.9022]

def clone_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font      = copy.copy(src_cell.font)
        dst_cell.fill      = copy.copy(src_cell.fill)
        dst_cell.border    = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)

# ---- Tambahkan header di baris 2 ----
header_row = 2
header_labels = ['Precision', 'Recall', 'Precision (%)', 'Recall (%)']
for idx, label in enumerate(header_labels):
    col = 7 + idx
    src  = ws.cell(row=header_row, column=3)
    dst  = ws.cell(row=header_row, column=col)
    dst.value = label
    clone_cell_style(src, dst)

# ---- Tambahkan data per kelas (baris 3-8) ----
for i, row_num in enumerate(range(3, 9)):
    p = precision_per_class[i]
    r = recall_per_class[i]

    # Precision (raw)
    c = ws.cell(row=row_num, column=7, value=round(p, 4))
    clone_cell_style(ws.cell(row=row_num, column=3), c)
    c.number_format = '0.0000'

    # Recall (raw)
    c = ws.cell(row=row_num, column=8, value=round(r, 4))
    clone_cell_style(ws.cell(row=row_num, column=3), c)
    c.number_format = '0.0000'

    # Precision (%)
    c = ws.cell(row=row_num, column=9, value=f'{p*100:.2f}%')
    clone_cell_style(ws.cell(row=row_num, column=5), c)

    # Recall (%)
    c = ws.cell(row=row_num, column=10, value=f'{r*100:.2f}%')
    clone_cell_style(ws.cell(row=row_num, column=5), c)

# Atur lebar kolom baru
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 14
ws.column_dimensions['J'].width = 14

wb.save(XLSX)
print('Excel berhasil diupdate dengan Precision & Recall per kelas!')
